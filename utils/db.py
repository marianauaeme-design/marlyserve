import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
from datetime import datetime, date
import streamlit as st

DB_PATH = "barbacoa_db.xlsx"

MENU_DEFAULT = {
    "Barbacoa (kg)": [
        ("1 Kilo barbacoa", 700),
        ("3/4 Kilo barbacoa", 525),
        ("1/2 Kilo barbacoa", 350),
        ("1/4 Kilo barbacoa", 175),
    ],
    "Tacos": [
        ("Taco panza / barbacoa", 35),
        ("Taco espinazo / pescuezo", 50),
        ("Taco pata / creadilla", 40),
    ],
    "Caldos": [
        ("Panza de res", 80),
        ("Menudo", 80),
        ("Consomé", 35),
        ("Consomé c/pata", 65),
    ],
    "Quesadillas": [
        ("Quesadilla champiñón / rajas", 45),
        ("Quesadilla queso c/barbacoa", 70),
    ],
    "Bebidas": [
        ("Atole", 30),
        ("Café", 30),
        ("Refresco", 25),
        ("Agua natural", 18),
        ("Agua de sabor", 30),
        ("Jarra de agua", 60),
    ],
    "Postres": [
        ("Gelatina", 25),
        ("Pan de amasijo", 25),
        ("Dulces", 17),
    ],
}

MESEROS_DEFAULT = [
    ("Ana", "1111"),
    ("Carlos", "2222"),
    ("Lupita", "3333"),
    ("Miguel", "4444"),
]

ADMIN_PIN = "0000"
NARANJA = "E85D04"
NARANJA_LT = "FAEEDA"
NARANJA_DK = "7C2D12"
HEADER_FILL = PatternFill("solid", fgColor=NARANJA)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)


def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def init_db():
    if os.path.exists(DB_PATH):
        return
    wb = Workbook()

    # --- Hoja: Menú ---
    ws = wb.active
    ws.title = "Menu"
    ws.append(["Categoria", "Producto", "Precio"])
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for cat, items in MENU_DEFAULT.items():
        for nombre, precio in items:
            ws.append([cat, nombre, precio])
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 10

    # --- Hoja: Meseros ---
    ws2 = wb.create_sheet("Meseros")
    ws2.append(["Nombre", "PIN", "Activo"])
    for cell in ws2[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for nombre, pin in MESEROS_DEFAULT:
        ws2.append([nombre, pin, True])
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 10

    # --- Hoja: Ordenes ---
    ws3 = wb.create_sheet("Ordenes")
    ws3.append(["ID", "Fecha", "Mesa", "Mesero", "Producto", "Precio_unit", "Cantidad", "Subtotal", "Estado", "Hora_apertura", "Hora_cierre", "Forma_pago", "Nota"])
    for cell in ws3[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for col in ["A","B","C","D","E","F","G","H","I","J","K"]:
        ws3.column_dimensions[col].width = 16

    # --- Hoja: Cortes ---
    ws4 = wb.create_sheet("Cortes")
    ws4.append(["Fecha", "Hora_corte", "Mesero", "Ordenes_cerradas", "Total_vendido", "Notas"])
    for cell in ws4[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    for col in ["A","B","C","D","E","F"]:
        ws4.column_dimensions[col].width = 18

    # --- Hoja: Config ---
    ws5 = wb.create_sheet("Config")
    ws5.append(["Clave", "Valor"])
    ws5.append(["admin_pin", ADMIN_PIN])
    ws5.append(["num_mesas", 15])
    ws5.append(["negocio", "Barbacoa Tío R"])
    ws5.column_dimensions["A"].width = 20
    ws5.column_dimensions["B"].width = 20

    wb.save(DB_PATH)


# ─── MENÚ ────────────────────────────────────────────────────────────────────

def get_menu():
    wb = load_workbook(DB_PATH, data_only=True)
    ws = wb["Menu"]
    menu = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        cat, prod, precio = row[0], row[1], row[2]
        if cat not in menu:
            menu[cat] = []
        menu[cat].append({"nombre": prod, "precio": int(precio or 0)})
    return menu


def update_menu(cat, nombre, precio_nuevo):
    wb = load_workbook(DB_PATH)
    ws = wb["Menu"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == cat and row[1].value == nombre:
            row[2].value = precio_nuevo
            break
    wb.save(DB_PATH)


def add_menu_item(cat, nombre, precio):
    wb = load_workbook(DB_PATH)
    ws = wb["Menu"]
    ws.append([cat, nombre, precio])
    wb.save(DB_PATH)


def delete_menu_item(cat, nombre):
    wb = load_workbook(DB_PATH)
    ws = wb["Menu"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == cat and row[1].value == nombre:
            ws.delete_rows(row[0].row)
            break
    wb.save(DB_PATH)


# ─── MESEROS ─────────────────────────────────────────────────────────────────

def get_meseros():
    wb = load_workbook(DB_PATH, data_only=True)
    ws = wb["Meseros"]
    meseros = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            meseros.append({"nombre": row[0], "pin": str(row[1]), "activo": row[2]})
    return meseros


def add_mesero(nombre, pin):
    wb = load_workbook(DB_PATH)
    ws = wb["Meseros"]
    ws.append([nombre, pin, True])
    wb.save(DB_PATH)


def update_pin(nombre, pin_nuevo):
    wb = load_workbook(DB_PATH)
    ws = wb["Meseros"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == nombre:
            row[1].value = pin_nuevo
            break
    wb.save(DB_PATH)


def toggle_mesero(nombre, activo):
    wb = load_workbook(DB_PATH)
    ws = wb["Meseros"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == nombre:
            row[2].value = activo
            break
    wb.save(DB_PATH)


def delete_mesero(nombre):
    wb = load_workbook(DB_PATH)
    ws = wb["Meseros"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == nombre:
            ws.delete_rows(row[0].row)
            break
    wb.save(DB_PATH)


def get_admin_pin():
    wb = load_workbook(DB_PATH, data_only=True)
    ws = wb["Config"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == "admin_pin":
            return str(row[1])
    return ADMIN_PIN


def get_num_mesas():
    wb = load_workbook(DB_PATH, data_only=True)
    ws = wb["Config"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == "num_mesas":
            return int(row[1])
    return 15


def update_config(clave, valor):
    wb = load_workbook(DB_PATH)
    ws = wb["Config"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == clave:
            row[1].value = valor
            wb.save(DB_PATH)
            return
    ws.append([clave, valor])
    wb.save(DB_PATH)


# ─── ÓRDENES ─────────────────────────────────────────────────────────────────

def get_next_id():
    wb = load_workbook(DB_PATH, data_only=True)
    ws = wb["Ordenes"]
    ids = [row[0] for row in ws.iter_rows(min_row=2, values_only=True) if row[0]]
    return max(ids) + 1 if ids else 1


def guardar_orden(mesa, mesero, items, hora_apertura, forma_pago="Efectivo"):
    wb = load_workbook(DB_PATH)
    ws = wb["Ordenes"]
    oid = get_next_id()
    hoy = date.today().isoformat()
    for item in items:
        subtotal = item["precio"] * item["cantidad"]
        nota = item.get("nota", "") or ""
        ws.append([
            oid, hoy, mesa, mesero,
            item["nombre"], item["precio"], item["cantidad"], subtotal,
            "abierta", hora_apertura, None, forma_pago, nota
        ])
    wb.save(DB_PATH)
    return oid


def cerrar_orden_db(mesa):
    wb = load_workbook(DB_PATH)
    ws = wb["Ordenes"]
    ahora = datetime.now().strftime("%H:%M")
    hoy = date.today().isoformat()
    for row in ws.iter_rows(min_row=2):
        if (str(row[2].value) == str(mesa)
                and row[8].value == "abierta"
                and str(row[1].value) == hoy):
            row[8].value = "cerrada"
            row[10].value = ahora
    wb.save(DB_PATH)


# ─── REPORTES ────────────────────────────────────────────────────────────────

def get_ventas_hoy():
    wb = load_workbook(DB_PATH, data_only=True)
    ws = wb["Ordenes"]
    hoy = date.today().isoformat()
    total = 0
    ordenes_ids = set()
    ventas_mesero = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[1]) == hoy and row[8] == "cerrada":
            total += row[7] or 0
            ordenes_ids.add(row[0])
            mesero = row[3] or "Sin asignar"
            ventas_mesero[mesero] = ventas_mesero.get(mesero, 0) + (row[7] or 0)
    return total, len(ordenes_ids), ventas_mesero


def get_ventas_rango(fecha_ini, fecha_fin):
    wb = load_workbook(DB_PATH, data_only=True)
    ws = wb["Ordenes"]
    fi, ff = str(fecha_ini), str(fecha_fin)
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] and fi <= str(row[1]) <= ff and row[8] == "cerrada":
            rows.append({
                "id": row[0], "fecha": row[1], "mesa": row[2],
                "mesero": row[3], "producto": row[4],
                "precio": row[5], "cantidad": row[6], "subtotal": row[7],
            })
    return rows


def registrar_corte(mesero_nombre, ordenes, total, notas=""):
    wb = load_workbook(DB_PATH)
    ws = wb["Cortes"]
    ahora = datetime.now()
    ws.append([
        ahora.date().isoformat(),
        ahora.strftime("%H:%M"),
        mesero_nombre,
        ordenes,
        total,
        notas,
    ])
    wb.save(DB_PATH)


def get_cortes():
    wb = load_workbook(DB_PATH, data_only=True)
    ws = wb["Cortes"]
    cortes = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            cortes.append({
                "fecha": row[0], "hora": row[1], "mesero": row[2],
                "ordenes": row[3], "total": row[4], "notas": row[5],
            })
    return list(reversed(cortes))
